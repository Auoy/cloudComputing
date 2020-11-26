# -*- conding:utf-8 -*-
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import openpyxl.styles as sty
from openpyxl import Workbook, load_workbook
import json


class Convert:

    @staticmethod
    def json2xls(fileName):
        wb1 = Workbook()
        sheet = wb1.active
        listHead = []
        data = json.load(open(fileName + ".json", "r", encoding="utf-8"))
        for c, i in enumerate(data[0].keys()):
            sheet.cell(row=1, column=c + 1, value=i)
            listHead.append(i)
        for r, i in enumerate(data):
            row = r + 2
            for c, d in enumerate(listHead):
                sheet.cell(row=row, column=c + 1, value=i.get(d, ""))

        wb1.save(fileName + ".xlsx")

    @staticmethod
    def xls2json(fileName, r, c):
        wb = load_workbook(fileName + ".xlsx")
        ws = wb["data"]
        list_key = []
        jsonLine = []
        for col in range(1, c + 1):
            list_key.append(ws.cell(row=1, column=col).value)
        for row in range(2, r + 1):
            dict_v = {}
            for col in range(1, c + 1):
                dict_v[list_key[col - 1]] = ws.cell(row=row, column=col).value
            jsonLine.append(dict_v)
        json.dump(jsonLine, open(fileName + ".json", "w", encoding="utf-8"), ensure_ascii=False)


if __name__ == "__main__":
    Convert.xls2json("result-2020-11-24-10-50-52", 1000, 7)  # 第二个参数是行数 第三个参数不需要改
