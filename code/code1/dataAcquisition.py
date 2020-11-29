import requests
import json
from openpyxl import Workbook
import time
import hashlib
import os
import datetime

from pymongo import MongoClient

import parameter

# 数据区
max_behot_time = '0'  # 链接参数
title = []  # 存储新闻标题
source_url = []  # 存储新闻的链接
s_url = []  # 存储新闻的完整链接
source = []  # 存储发布新闻的公众号
media_url = {}  # 存储公众号的完整链接
time_table = []  # 时间戳
time_change = []  # 转化后的标准时间的时间戳
chinese_tag = []  # 新闻类型


def get_as_cp():  # 该函数主要是为了获取as和cp参数，程序参考今日头条中的加密js文件：home_4abea46.js
    zz = {}
    now = round(time.time())
    # print(now) # 获取当前计算机时间
    e = hex(int(now)).upper()[2:]  # hex()转换一个整数对象为16进制的字符串表示
    # print('e:', e)
    a = hashlib.md5()  # hashlib.md5().hexdigest()创建hash对象并返回16进制结果
    # print('a:', a)
    a.update(str(int(now)).encode('utf-8'))
    i = a.hexdigest().upper()
    # print('i:', i)
    if len(e) != 8:
        zz = {'as': '479BB4B7254C150', 'cp': '7E0AC8874BB0985'}
        return zz
    n = i[:5]
    a = i[-5:]
    r = ''
    s = ''
    for i in range(5):
        s = s + n[i] + e[i]
    for j in range(5):
        r = r + e[j + 3] + a[j]
    zz = {'as': 'A1' + s + e[-3:], 'cp': e[0:3] + r + 'E1'}
    # print('zz:', zz)
    return zz


def getdata(url, headers, cookies):  # 解析网页函数
    r = requests.get(url, headers=headers, cookies=cookies)
    print(url)
    data = json.loads(r.text)
    print(r.text)
    return data


def savedata(title, s_url, source, media_url, time_table, time_change, chinese_tag):  # 存储数据到文件
    # 存储数据到xlxs文件
    wb = Workbook()
    if not os.path.isdir(os.getcwd() + '/result'):  # 判断文件夹是否存在
        os.makedirs(os.getcwd() + '/result')  # 新建存储文件夹
    filename = os.getcwd() + '/result/result-' + datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S') + '.xlsx'
    # 新建存储结果的excel文件
    ws = wb.active
    ws.title = 'data'  # 更改工作表的标题
    ws['A1'] = '标题'  # 对表格加入标题
    ws['B1'] = '新闻链接'
    ws['C1'] = '头条号'
    ws['D1'] = '头条号链接'
    ws['E1'] = '时间戳'
    ws['F1'] = '标准时间'
    ws['G1'] = '新闻类型'
    for row in range(2, len(title) + 2):  # 将数据写入表格
        _ = ws.cell(column=1, row=row, value=title[row - 2])
        _ = ws.cell(column=2, row=row, value=s_url[row - 2])
        _ = ws.cell(column=3, row=row, value=source[row - 2])
        _ = ws.cell(column=4, row=row, value=media_url[source[row - 2]])
        _ = ws.cell(column=5, row=row, value=time_table[row - 2])
        _ = ws.cell(column=6, row=row, value=time_change[row - 2])
        _ = ws.cell(column=7, row=row, value=chinese_tag[row - 2])
    wb.save(filename=filename)  # 保存文件


def main(max_behot_time, title, source_url, s_url, source, media_url, chinese_tag):  # 主函数
    for i in range(20):  # 此处的数字类似于你刷新新闻的次数，正常情况下刷新一次会出现10条新闻，但夜存在少于10条的情况；所以最后的结果并不一定是10的倍数
        ascp = get_as_cp()  # 获取as和cp参数的函数
        demo = getdata(
            parameter.start_url + max_behot_time + '&max_behot_time_tmp=' + max_behot_time + '&tadrequire=true&as=' +
            ascp['as'] + '&cp=' + ascp['cp'],
            parameter.headers, parameter.cookies)
        print(demo)
        time.sleep(1)
        for j in range(len(demo['data'])):
            # print(demo['data'][j]['title'])
            if demo['data'][j]['title'] not in title:
                title.append(demo['data'][j]['title'])  # 获取新闻标题
                source_url.append(demo['data'][j]['source_url'])  # 获取新闻链接
                source.append(demo['data'][j]['source'])  # 获取发布新闻的公众号
                time_table.append(demo['data'][j]['behot_time'])
                timeArray = time.localtime(demo['data'][j]['behot_time'])
                otherTimeStyle = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
                time_change.append(otherTimeStyle)
                if ("chinese_tag" in demo['data'][j]):
                    chinese_tag.append(demo['data'][j]['chinese_tag'])
                else:
                    chinese_tag.append("unknown")
            if demo['data'][j]['source'] not in media_url:
                media_url[demo['data'][j]['source']] = parameter.url + demo['data'][j]['media_url']  # 获取公众号链接
        print(max_behot_time)
        max_behot_time = str(demo['next']['max_behot_time'])  # 获取下一个链接的max_behot_time参数的值
        for index in range(len(title)):
            print('标题：', title[index])
            if 'https' not in source_url[index]:
                s_url.append(parameter.url + source_url[index])
                print('新闻链接：', parameter.url + source_url[index])
            else:
                print('新闻链接：', source_url[index])
                s_url.append(source_url[index])
            # print('源链接：', url+source_url[index])
            print('头条号：', source[index])
            print('时间戳', time_change[index])
            print('新闻类型', chinese_tag[index])
            print(len(title))  # 获取的新闻数量


def convertJson():
    if not os.path.isdir(os.getcwd() + '/json'):  # 判断文件夹是否存在
        os.makedirs(os.getcwd() + '/json')  # 新建存储文件夹
    shu = {}
    for index in range(len(title)):
        collect = {}
        collect['title'] = title[index]
        collect['url'] = source_url[index]
        collect['source'] = source[index]
        collect['media_url'] = media_url[source[index]]
        collect['time'] = time_change[index]
        collect['tags'] = chinese_tag[index]
        shu[index] = collect
    json_str = json.dumps(shu, ensure_ascii=False)
    s = os.getcwd() + "/json/result-" + datetime.datetime.now().strftime('%Y-%m-%d-%H-%M-%S') + '.json'
    host = '127.0.0.1'  # 本机ip地址
    client = MongoClient(host, 27017)  # 建立客户端对象
    write_to_mongodb(client, json_str)
    with open(s, 'w', encoding='utf-8') as json_file:
        json_file.write(json_str)


# 将读到的标题数据写入 MongoDB 中，在 convertJson() 方法中调用
def write_to_mongodb(client, data):
    db = client.mydb  # 连接mydb数据库，没有则自动创建
    myset = db.testset  # 使用test_set集合，没有则自动创建
    for val in eval(data).values():
        myset.insert_one(val)


if __name__ == '__main__':
    main(max_behot_time, title, source_url, s_url, source, media_url, chinese_tag)
    savedata(title, s_url, source, media_url, time_table, time_change, chinese_tag)
    convertJson()
